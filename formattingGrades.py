# John Everett, Kiera Fisher,, Aysha Entrikin, Meagan Brown, Will Francom, Zaniel Murdock
# Section 004
# Description: A program that automatically formats and summarizes the important
# information about each of the classes they teach

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import statistics

# Load the existing workbook
myWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
sourceSheet = myWorkbook.active

# Function to create a worksheet for a specific course
def createWorksheet(course):
    # Create a new sheet in the existing workbook
    newSheet = myWorkbook.create_sheet(title=course)

    # Headers
    headers = ["Last Name", "First Name", "Student ID", "Grade", "", "Summary Statistics", "Value"]
    for col_num, header in enumerate(headers, start=1):
        newSheet.cell(row=1, column=col_num, value=header).font = Font(bold= True)

    # Adjust column widths for readability
    for col in ["A", "B", "C", "D", "F", "G"]:
        cell_value = newSheet[f"{col}1"].value
        newSheet.column_dimensions[col].width = len(cell_value) + 5

    stats = ["Highest Grade", "Lowest Grade", "Mean Grade", "Median Grade", "Number of Students"]
    for row, stat in enumerate(stats, start = 2) :
        newSheet.cell(row = row, column = 6, value = stat)

    row_index = 2  # Start inserting data from row 2
    grades = []

    # Loop through rows to find students in the specified course
    for row in range(2, sourceSheet.max_row + 1):
        cell_value = sourceSheet[f"A{row}"].value  # Course name
        
        if cell_value == course:
            student_data = sourceSheet[f"B{row}"].value  # Student data format: "Lucy_Jane_001001"

            if student_data:
                parts = student_data.split("_")
                if len(parts) == 3:
                    last_name, first_name, student_id = parts[0], parts[1], parts[2]

                    newSheet[f"A{row_index}"] = last_name
                    newSheet[f"B{row_index}"] = first_name
                    newSheet[f"C{row_index}"] = student_id
                    newSheet[f"D{row_index}"] = sourceSheet[f"C{row}"].value  # Grade column
                    row_index += 1  # Move to the next row
            grade = sourceSheet[f"C{row}"].value
            grades.append(grade)
    
    # Calculate course statistics
    max_grade = max(grades)
    min_grade = min(grades)
    mean_grade = sum(grades) / len(grades)
    median_grade = statistics.median(grades)

    # Print course statistics
    newSheet["G2"] = max_grade
    newSheet["G3"] = min_grade
    newSheet["G4"] = mean_grade
    newSheet["G5"] = median_grade
    newSheet["G6"] = len(grades)

    # Apply autofilter
    newSheet.auto_filter.ref = f"A1:D{row_index}"

# Create sheets for different courses
for course_name in ["Algebra", "Trigonometry", "Geometry", "Calculus", "Statistics"]:
    createWorksheet(course_name)

# Remove uncleaned data worksheet
myWorkbook.remove(sourceSheet)

# Save and close the workbook
myWorkbook.save("formatted_grades.xlsx")
myWorkbook.close()

print("Done")
